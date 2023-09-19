from django.shortcuts import render
from django.conf import settings
from django.db.models import Max
from django.core.files.base import ContentFile
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .serializer import workOrderSerializer, aiWorkOrderSerializer, OrderSerializer
from .models import workOrder, aiWorkOrder, Order, OrderItem

from .robot import main, robot_control, speed
# from .main_result_20230830 import activate_cal
from .main_result_20230911 import activate_cal
from io import BytesIO
import random
import qrcode
import pandas as pd
import time
import uuid
import csv
import os


@api_view(['POST'])
def controlRobot(request):
    data = request.data
    csv_id = data.get('id')
    order = workOrder.objects.filter(id=csv_id).first()
    print(data)
    # print(order)
    # print('order total:',order.total_count)
    robot_speed = int(data.get('speed')) if int(data.get('speed')) <= 100 else 100
    txt_path = os.path.join(settings.MEDIA_ROOT, 'output.txt')

    # if data.get('mode') == 'activate':
    #     main(csv_id, order.total_count)
    #     speed(50)
    # elif data.get('mode') == 'pause':
    #     robot_control('192.168.1.15', 10040).pause()
    # elif data.get('mode') == 're-activate':
    #     robot_control('192.168.1.15', 10040).start()
    # elif data.get('mode') == 'speed':
    #     speed(robot_speed)
    # elif data.get('mode') == 'reset':
    #     with open (txt_path, "w") as f:
    #         f.write(f'')
    #     robot_control('192.168.1.15',10040).pause()
    #     robot_control('192.168.1.15',10040).reset()

    # 1,準備抓取第1個物件,18,activate
    ai = aiWorkOrder.objects.filter(worklist_id=csv_id).first().list_order.split(',')

    if data.get('mode') == 'activate':
        time.sleep(2)
        for i, data in enumerate(ai):
            with open(txt_path, 'w', encoding='utf-8') as f:
                f.write(f'{i+1},準備抓取第{i+1}個物件,{data},prepare,1')
            time.sleep(3)
            with open(txt_path, 'w', encoding='utf-8') as f:
                f.write(f'{i+1},正在操作第{i+1}個物件,{data},operate,1')
            time.sleep(3)
        with open (txt_path, "w", encoding='utf-8') as f:
            f.write(f'')
        print("inner activate")
    elif data.get('mode') == 'pause':
        print("inner pause")
    elif data.get('mode') == 're-activate':
        print("inner re-activate")
    elif data.get('mode') == 'speed':
        print("inner speed")
    elif data.get('mode') == 'reset':
        with open (txt_path, "w") as f:
            f.write(f'')
        print("inner reset")

    # if data.get('mode') == 'activate':
    #     file = os.path.join(settings.MEDIA_ROOT, 'csv', 'box_bestPositions_solution_9_convey_20230829.csv')

    #     Supply = pd.read_csv(file)
    #     name = Supply['matched_box_name'].tolist()
    #     name = [i.replace("#", "").replace("外箱", "") for i in name]
    #     txt_path = os.path.join(settings.MEDIA_ROOT, "output.txt")
    #     with open(txt_path, 'w') as t:
    #         t.write(f'{0},{name[0]}')
    return Response({"robot"})

@api_view(['GET'])
def readTxt(request):
    txt_path = os.path.join(settings.MEDIA_ROOT, 'output.txt')
    with open(txt_path, 'r', encoding='utf-8') as f:
        content = f.read()
    
    return Response(content)

@api_view(['POST'])
def createWorkOrder(request):
    data = request.data
    
    csv_data = [
        ["box_id", "name", "width", "height", "depth", "quantity"],
        [1, "#7A外箱", 35, 26, 20, data.get("7A")],
        [2, "#9外箱", 43, 32, 23, data.get("9")],
        [3, "#16A外箱", 35, 26, 16, data.get("16A")],
        [4, "#18A外箱", 35, 26, 18, data.get("18A")],
        [5, "#13外箱", 56, 25, 14, data.get("13")],
        [6, "#20外箱", 53, 34, 13, data.get("20")],
        [7, "#22外箱", 45, 26, 18, data.get("22")],
        [8, "#26外箱", 72, 25, 20, data.get("26")],
        [9, "#29外箱", 65, 25, 18, data.get("29")],
        [10, "#33外箱", 44, 21, 18, data.get("33")],
        [11, "#35外箱", 102, 46, 18, data.get("35")],
    ]

    workOrder.objects.create(
        name = data.get("name"),
        _16A_qty = data.get("16A"),
        _18A_qty = data.get("18A"),
        _33_qty = data.get("33"),
        _7A_qty = data.get("7A"),
        _13_qty = data.get("13"),
        _22_qty = data.get("22"),
        _20_qty = data.get("20"),
        _29_qty = data.get("29"),
        _9_qty = data.get("9"),
        _26_qty = data.get("26"),
        _35_qty = data.get("35"),
    )

    max_id = workOrder.objects.aggregate(Max('id')).get('id__max')
    order = workOrder.objects.filter(id=max_id).first()
    order.file_name =  f'box_data_{max_id}.csv'
    order.save()

    csv_file = os.path.join(settings.MEDIA_ROOT, f'box_data_{max_id}.csv')
    with open(csv_file, 'w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
    
        for row in csv_data:
            writer.writerow(row)

    # 亂數
    # items = []
    # for key, value in data.items():
    #     items.extend([key] * value)
    # random.shuffle(items)
    # print("original data:", data)
    # print("random data:", items)
    return Response({'ok'})

@api_view(['GET'])
def getWorkOrderData(request):
    order = workOrder.objects.all().order_by('-id')
    serializer = workOrderSerializer(order, many=True)
    return Response(serializer.data)

@api_view(['POST'])
def aiCalculate(request):
    data = request.data
    worklist_id = data.get("id")
    t1 = time.time()
    activate_cal(worklist_id)
    t2 = time.time()
    training_time = round(t2-t1, 3)
    ai_csvfile_path = os.path.join(settings.MEDIA_ROOT, f'Figures_{worklist_id}', f'box_positions_final.csv')
    ai_df = pd.read_csv(ai_csvfile_path)
    ai_list = ai_df['matched_box_name'].tolist()
    ai_str = ','.join([ai.replace('#', '').replace('外箱', '') for ai in ai_list])
    work_order = workOrder.objects.filter(id=int(worklist_id)).first()
    work_order.hasAiTrained = True
    work_order.save()
    aiWorkOrder.objects.create(
        name = work_order.name,
        worklist_id = worklist_id,
        list_order = ai_str,
        training_time = training_time
    )

    return Response({"worklist_id": worklist_id, "list_order": ai_str, "training_time": training_time})
    # return Response({"worklist_id": worklist_id,  "training_time": training_time})

@api_view(['GET'])
def getAiWorkOrderData(request):
    ai_order = aiWorkOrder.objects.all().order_by('-id')
    serializer = aiWorkOrderSerializer(ai_order, many=True)
    return Response(serializer.data)

@api_view(['POST'])
def uploadCsv(request):
    # if 'csv_file' not in request.data:
    #     return Response({'message': '請提供有效的 CSV 檔案'})
    csv_file_length = int(request.data.get('csv_file_length'))

    try:
        for i in range(csv_file_length):
            csv_file = request.data.get(f'csv_file{i+1}')
            csv_name = request.data.get(f'csv_file_name{i+1}').replace('.csv', '')
            unique_code = uuid.uuid4().hex
            unique_code_exist = Order.objects.filter(unique_code=unique_code)

            while unique_code_exist:
                unique_code = uuid.uuid4().hex
                unique_code_exist = Order.objects.filter(unique_code=unique_code)
                if unique_code_exist is None:
                    break
            
            # image = qrcode.make(unique_code)
            # # 将图像保存到模型的ImageField字段
            # image_io = BytesIO()
            # image.resize((150, 150))
            # image.save(image_io, 'PNG')
            # image_content = ContentFile(image_io.getvalue(), name=unique_code + '.png')

            qr = qrcode.QRCode(
                version=1,  # 版本（1-40），版本越高，QR码容纳的数据越多
                error_correction=qrcode.constants.ERROR_CORRECT_L,  # 错误纠正级别
                box_size=10,  # 每个模块的像素大小
                border=4,  # 边框大小
            )

            qr.add_data(unique_code)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            img = img.resize((120, 120))  # 指定大小为150x150
            image_io = BytesIO()
            img.save(image_io, 'PNG')
            image_content = ContentFile(image_io.getvalue(), name=unique_code + '.png')

            order = Order.objects.create(
                name = csv_name,
                unique_code = unique_code,
                image = image_content)
            order.save()
            
            # 使用 csv.reader 來讀取 CSV 檔案內容
            csv_reader = csv.reader(csv_file.read().decode('utf-8').splitlines())
            next(csv_reader)
            for reader in csv_reader:
                OrderItem.objects.create(
                    order = order,
                    name = reader[1].replace('外箱', ''),
                    width = reader[2],
                    height = reader[3],
                    depth = reader[4],
                    count = int(reader[5])
                )

            # xlsx
            # xlsx_file = request.FILES.get(f'csv_file{i+1}')
            # df = pd.read_excel(xlsx_file)
            # print(df.to_dict())

            

    except Exception as e:
        return Response({'message': 'error'})

    return Response({'message': 'CSV 檔案解析成功', 'data': 2}, status=200)

@api_view(['GET'])
def getOrderData(request):
    order = Order.objects.all().order_by('-id')
    serializer = OrderSerializer(order, many=True)
    return Response(serializer.data)
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment

@api_view(['GET'])
def getOrderXlsxFile(request, id):
    print(id)
    xlsx_path = os.path.join(settings.MEDIA_ROOT, 'order.xlsx')
    # df = pd.read_excel(xlsx_path)
    # print(df.to_dict)

    # data = {'怡振電機-AI技術部混料堆疊工單': {0: '序號', 1: 1, 2: 2, 3: 3}, 'Unnamed: 1': {0: '訂單名稱', 1: 'box_data_57', 2: 'box_data_58', 3: 'box_data_59'}, 'Unnamed: 2': {0: 'asd', 1: 'asd', 2: 'asd', 3: 'asd'}, 'Unnamed: 3': {0: '訂單內容', 1: '1.#7A_5\n2.#9_4\n3.#16A_5\n4.#18A_5\n5.#13_5\n6.#20_5\n7.#22_5\n8.#26_5\n9.#29_5\n10.#33_4\n11.#35_3', 2: '1.#16A_1\n2.#33_1', 3: '1.#18A_1'}, 'Unnamed: 4': {0: 'asd', 1: 'asd', 2: 'asd', 3: 'asd'}, 'Unnamed: 5': {0: 'asd', 1: 'asd', 2: 'asd', 3: 'asd'}, 'Unnamed: 6': {0: '時間', 1: '2023年9月15日\n下午 05:09:42', 2: '2023年9月15日\n下午 05:10:42', 3: '2023年9月15日\n下午 05:11:42'}, 'Unnamed: 7': {0: '工單QR-Code', 1: 'asd', 2: 'asd', 3: 'asd'}}

    # df = pd.DataFrame(data)

    # # 将数据框写入.xlsx文件
    # df.to_excel(os.path.join(settings.MEDIA_ROOT, 'test.xlsx'), index=False)

   

    # 打开工作簿
    workbook = openpyxl.load_workbook(xlsx_path)

    # 选择工作表
    sheet = workbook.active  # 或者通过工作表名称：sheet = workbook['Sheet1']

    # 读取数据
    # for row in sheet.iter_rows(values_only=True):
    #     print(row)
    #     for cell in row:
    #         print(cell)

    # 添加新行
    data_to_add = (4, 'box_data_59', None, '1.#18A_1', None, None, '2023年9月15日\n下午 05:11:42')
    sheet['A6'] = 4
    sheet['A6'].alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells('B6:C6')
    sheet['B6'] = 'box_data_59'
    sheet['B6'].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    sheet.merge_cells('D6:F6')
    text = '1.#7A_5\n2.#9_4\n3.#16A_5\n4.#18A_5\n5.#13_5'
    sheet['D6'] = text
    sheet['D6'].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    sheet['G6'] = '2023年9月15日\n下午 05:11:42'
    sheet['G6'].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    text_lines = text.split('\n')  # 如果文本中有换行符，请按换行符拆分文本
    max_line_length = max(len(line) for line in text_lines)
    row_height = 15  # 自定义行高
    if max_line_length > 0:
        row_height *= (len(text_lines) + 1)  # 增加行数以适应文本
    sheet.row_dimensions[6].height = row_height + 30

    sheet.merge_cells('H6:I6')
    # text = '1.#7A_5\n2.#9_4\n3.#16A_5\n4.#18A_5\n5.#13_5'
    # sheet['H6'] = text

    # 加载图片文件
    img = Image(os.path.join(settings.MEDIA_ROOT, 'qrcode', '3873dfe0db444f9c987eb669e1f84337.png'))  # 替换为实际图片路径
    
    # 设置图片的位置和大小

    sheet.add_image(img, 'H6')  # 图片跨越从A1到B2的单元格
    # 设置H6:I6单元格的水平和垂直居中对齐
    # for row in sheet.iter_rows(min_row=6, max_row=6, min_col=8, max_col=9):
    #     for cell in row:
    #         cell.alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    sheet['H6'].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
    sheet.row_dimensions[6].height = 11*15 + 30
    # 保存工作簿
    workbook.save(os.path.join(settings.MEDIA_ROOT, 'new.xlsx'))
    return Response({'ok'})

